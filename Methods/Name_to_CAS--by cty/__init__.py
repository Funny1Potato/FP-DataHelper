import os
import openpyxl
import time
import sys
import json
import re
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from collections import Counter

# ================= 复用之前的识别函数 =================
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]

TIME_KEYWORDS = ('保留时间', '时间', 'time')
NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')

def is_time_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in TIME_KEYWORDS)

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in NAME_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    """返回同时包含时间列和名称列的行索引（0-based）。"""
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r+1]]
        has_time = any(is_time_header(v) for v in row_vals)
        has_name = any(is_name_header(v) for v in row_vals)
        if has_time and has_name:
            return r
    return 0

# ================= 参考程序中的 CAS 查询函数（完全照抄） =================
def get_cas_by_name(name):
    """
    根据物质名称去 NIST 搜索 CAS 号。
    """
    if name is None or str(name).strip() == "":
        return "\\"
    
    clean_name = str(name).strip()
    url = "https://webbook.nist.gov/cgi/cbook.cgi"
    params = {
        "Name": clean_name,
        "Units": "SI"
    }
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }

    try:
        response = requests.get(url, params=params, headers=headers, timeout=15)
        
        if response.status_code != 200:
            return "Connection Error"
            
        soup = BeautifulSoup(response.text, 'html.parser')
        
        if "Search Results" in soup.title.text:
            return "Ambiguous/List Found"
            
        if "Name Not Found" in response.text:
            return "Not Found"

        cas_label = soup.find('strong', string="CAS Registry Number:")
        if cas_label:
            cas_text = cas_label.next_sibling
            if cas_text:
                return cas_text.strip()
        
        all_text = soup.get_text()
        if "CAS Registry Number:" in all_text:
            match = re.search(r'CAS Registry Number:\s*([\d\-]+)', all_text)
            if match:
                return match.group(1)

        return "Not Found in Page"

    except Exception as e:
        print(f"Error processing {name}: {e}")
        return "Error"

# ================= 处理单个工作表，添加 CAS 列 =================
def process_sheet_add_cas(ws, name_col_param, show_progress=True):
    """
    读取工作表，在名称列前插入 CAS 列，返回 (上方行, 新表头, 新数据行)
    """
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row+1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
        rows.append(row_vals)

    # 检测表头行
    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx+1:]

    # 确定名称列索引
    if name_col_param != "":
        # 将字母转为索引（0-based）
        name_col = HEAD.index(name_col_param.upper())
    else:
        # 自动识别：在表头行中找名称列
        name_col = None
        for i, val in enumerate(header_row_vals):
            if is_name_header(val):
                name_col = i
                break
        if name_col is None:
            # 如果找不到，默认第一列
            name_col = 0

    # 收集所有需要查询的名称（去重）
    name_set = set()
    for row in data_rows:
        if name_col < len(row):
            val = row[name_col]
            if val is not None and str(val).strip():
                name_set.add(str(val).strip())

    # 查询 CAS，建立缓存
    cas_cache = {}
    desc = "查询 CAS 号"
    for name in tqdm(name_set, desc=desc, disable=not show_progress, unit="个"):
        cas = get_cas_by_name(name)
        cas_cache[name] = cas
        time.sleep(1)   # 礼貌爬虫

    # 构建新表头：在名称列前插入 "CAS号"
    new_header = list(header_row_vals)
    new_header.insert(name_col, "CAS号")

    # 构建新数据行
    new_data_rows = []
    for row in data_rows:
        # 获取名称
        name_val = row[name_col] if name_col < len(row) else None
        if name_val is None or str(name_val).strip() == "":
            cas_val = ""
        else:
            cas_val = cas_cache.get(str(name_val).strip(), "Not Found")
        # 插入 CAS
        new_row = list(row)
        new_row.insert(name_col, cas_val)
        new_data_rows.append(new_row)

    return above_rows, new_header, new_data_rows

# ================= 主程序 =================
def main():
    option = json.loads(sys.argv[1])
    file = option["File"]
    name_col_param = option.get("Para1")  # 如 "A"，可选

    # 加载工作簿
    wb = openpyxl.load_workbook(file)
    out_wb = openpyxl.Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
        ws = wb[sheet_name]
        above_rows, new_header, new_data_rows = process_sheet_add_cas(ws, name_col_param, show_progress=True)

        out_ws = out_wb.create_sheet(title=sheet_name)

        # 写入上方行（不插入 CAS）
        for r_idx, row_vals in enumerate(above_rows):
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=r_idx+1, column=c_idx+1, value=val)

        # 写入新表头
        header_row_num = len(above_rows) + 1
        for c_idx, val in enumerate(new_header):
            if val is not None:
                out_ws.cell(row=header_row_num, column=c_idx+1, value=val)

        # 写入数据行
        for r_idx, row_vals in enumerate(new_data_rows):
            row_num = header_row_num + 1 + r_idx
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=row_num, column=c_idx+1, value=val)

    # 保存
    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}_with_cas.xlsx")
    out_wb.save(out_path)
    print(f"处理完成，文件保存至：{out_path}")

if __name__ == "__main__":
    main()