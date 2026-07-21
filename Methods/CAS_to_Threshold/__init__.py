import os
import openpyxl
import time
import sys
import json
from tqdm import tqdm

# ================= 识别函数 =================
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]

TIME_KEYWORDS = ('保留时间', '时间', 'time')
NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')
CAS_KEYWORDS = ('cas', 'cas号', 'cas number')

def is_time_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in TIME_KEYWORDS)

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in NAME_KEYWORDS)

def is_cas_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in CAS_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    """返回同时包含时间列和名称列的行索引（0-based）。"""
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r+1]]
        has_time = any(is_time_header(v) for v in row_vals)
        has_name = any(is_name_header(v) for v in row_vals)
        if has_time and has_name:
            return r
    return 0

# ================= 阈值加载 =================
def load_thresholds(json_path="./Database/Threshold.json"):
    """从 JSON 文件读取阈值数据，返回字典 {cas: threshold}"""
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"未找到阈值文件: {json_path} ")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

# ================= 处理单个工作表，添加“嗅觉阈值”列 =================
def process_sheet_add_threshold(ws, cas_col_param, threshold_dict, show_progress=True):
    """
    在 CAS 列后插入“嗅觉阈值”列，从 threshold_dict 中匹配。
    """
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row+1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
        rows.append(row_vals)

    # 检测表头行
    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx+1:]

    # 确定 CAS 列索引
    if cas_col_param != "":
        cas_col = HEAD.index(cas_col_param.upper())
    else:
        cas_col = None
        for i, val in enumerate(header_row_vals):
            if is_cas_header(val):
                cas_col = i
                break
        if cas_col is None:
            raise ValueError(f"工作表 '{ws.title}' 中未找到 CAS 列，且未指定列字母。")

    # 构建新表头：在 CAS 列后插入“嗅觉阈值”
    new_header = list(header_row_vals)
    new_header.insert(cas_col + 1, "嗅觉阈值")

    # 构建新数据行
    new_data_rows = []
    for row in tqdm(data_rows, desc=f"处理 {ws.title} 数据行", disable=not show_progress, unit="行"):
        cas_val = row[cas_col] if cas_col < len(row) else None
        if cas_val is None or str(cas_val).strip() == "":
            threshold = ""
        else:
            cas_key = str(cas_val).strip()
            threshold = threshold_dict.get(cas_key, "Not Found")
        new_row = list(row)
        new_row.insert(cas_col + 1, threshold)
        new_data_rows.append(new_row)

    return above_rows, new_header, new_data_rows

# ================= 主程序 =================
def main():
    option = json.loads(sys.argv[1])
    file = option["File"]
    cas_col_param = option.get("Para1")  # 如 "D"

    # 加载阈值 JSON（使用相对路径）
    try:
        threshold_dict = load_thresholds("./Database/Threshold.json")
        print("已加载阈值数据文件")
    except FileNotFoundError as e:
        print(f"错误: {e}")
        return

    wb = openpyxl.load_workbook(file)
    out_wb = openpyxl.Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
        ws = wb[sheet_name]
        try:
            above_rows, new_header, new_data_rows = process_sheet_add_threshold(
                ws, cas_col_param, threshold_dict, show_progress=True
            )
        except ValueError as e:
            print(f"工作表 '{sheet_name}' 处理失败: {e}")
            continue

        out_ws = out_wb.create_sheet(title=sheet_name)

        # 写入上方行
        for r_idx, row_vals in enumerate(above_rows):
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=r_idx+1, column=c_idx+1, value=val)

        # 写入新表头
        header_row_num = len(above_rows) + 1
        for c_idx, val in enumerate(new_header):
            if val is not None:
                out_ws.cell(row=header_row_num, column=c_idx+1, value=val)

        # 写入数据行
        for r_idx, row_vals in enumerate(new_data_rows):
            row_num = header_row_num + 1 + r_idx
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=row_num, column=c_idx+1, value=val)

    # 保存
    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}_with_threshold.xlsx")
    out_wb.save(out_path)
    print(f"处理完成，文件保存至：{out_path}")

if __name__ == "__main__":
    main()