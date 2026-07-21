import os
import openpyxl
import time
import sys
import json
from collections import defaultdict, Counter
import re
from tqdm import tqdm   # 导入 tqdm

# 列字母生成器（最多支持 3 个字母）
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]   # 足够覆盖常见列数

# 关键词（不区分大小写）
TIME_KEYWORDS = ('保留时间', '时间', 'time')
NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')

def is_time_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in TIME_KEYWORDS)

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in NAME_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    """
    扫描前 max_rows 行，返回同时包含时间列和名称列的行索引（0-based）。
    若未找到，返回 0。
    """
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r+1]]   # openpyxl 行号从 1 开始
        has_time = any(is_time_header(v) for v in row_vals)
        has_name = any(is_name_header(v) for v in row_vals)
        if has_time and has_name:
            return r
    return 0

def detect_groups(header_row_vals, count_param=None):
    """
    自动识别分组，返回 groups 和 count。
    若识别失败且未提供 count_param，则抛出 ValueError。
    """
    total_cols = len(header_row_vals)
    time_idxs = [i for i, v in enumerate(header_row_vals) if is_time_header(v)]
    name_idxs = [i for i, v in enumerate(header_row_vals) if is_name_header(v)]

    if not time_idxs:
        raise ValueError("未找到任何时间列（表头包含'保留时间/时间/time'），无法自动分组。")

    # 确定每组列数 count
    if len(time_idxs) >= 2:
        gaps = [time_idxs[i+1] - time_idxs[i] for i in range(len(time_idxs)-1)]
        counter = Counter(gaps)
        count = counter.most_common(1)[0][0]
    else:
        # 只有一个时间列，必须手动指定 count
        if count_param is None:
            raise ValueError("仅识别到单个时间列，无法自动确定分组列数，请指定每组列数。")
        count = count_param

    # 尝试确定时间列在组内的偏移 offset（0-based），使得每组恰好包含一个名称列且偏移一致
    possible_offsets = []
    for offset in range(count):
        start0 = time_idxs[0] - offset
        if start0 < 0 or start0 + count > total_cols:
            continue
        # 检查每个时间列所在组
        valid = True
        name_rel_offset = None  # 名称列在组内的相对偏移（若存在）
        for t_idx in time_idxs:
            start = t_idx - offset
            if start < 0 or start + count > total_cols:
                valid = False
                break
            # 本组内的所有列索引
            group_cols = list(range(start, start + count))
            # 找本组内的名称列
            group_name_idxs = [c for c in group_cols if c in name_idxs]
            if len(group_name_idxs) != 1:
                valid = False
                break
            # 计算名称列相对偏移
            rel = group_name_idxs[0] - start
            if name_rel_offset is None:
                name_rel_offset = rel
            elif name_rel_offset != rel:
                valid = False
                break
        if valid:
            possible_offsets.append((offset, name_rel_offset))

    if not possible_offsets:
        raise ValueError("无法根据时间列和名称列确定有效的分组，请检查表头列排列或手动指定每组列数。")

    # 取第一个可行的 offset（通常唯一）
    offset, name_rel_offset = possible_offsets[0]

    # 构建 groups
    groups = []
    for t_idx in time_idxs:
        start = t_idx - offset
        group_cols = list(range(start, start + count))
        time_col = t_idx
        name_col = start + name_rel_offset
        other_cols = [c for c in group_cols if c != time_col and c != name_col]
        groups.append({
            'time_col': time_col,
            'name_col': name_col,
            'other_cols': other_cols,
            'start': start,
            'count': count
        })
    return groups, count

def process_sheet(ws, dev, count_param, show_progress=True):
    """
    处理单个工作表，返回新表头、上方行、数据行。
    """
    # 1. 获取所有行值（保留 None）
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row+1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
        rows.append(row_vals)

    # 2. 检测表头行
    header_row_idx = detect_header_row(ws)
    header_row_vals = rows[header_row_idx] if header_row_idx < len(rows) else []
    # 上方行（0 ~ header_row_idx-1）
    above_rows = rows[:header_row_idx]

    # 3. 识别分组
    groups, count = detect_groups(header_row_vals, count_param)
    if not groups:
        # 若仍失败，强制默认单组（列数为 count）
        count = count_param if count_param is not None else 4
        total_cols = len(header_row_vals)
        groups = []
        for start in range(0, total_cols, count):
            if start + count > total_cols:
                break
            group_cols = list(range(start, start+count))
            # 尝试找名称列
            name_col = None
            for c in group_cols:
                if is_name_header(header_row_vals[c]):
                    name_col = c
                    break
            if name_col is None:
                name_col = group_cols[1] if len(group_cols)>1 else group_cols[0]
            time_col = group_cols[0]  # 默认第一列为时间
            other_cols = [c for c in group_cols if c != time_col and c != name_col]
            groups.append({
                'time_col': time_col,
                'name_col': name_col,
                'other_cols': other_cols,
                'start': start,
                'count': count
            })

    # 4. 提取数据（从表头行下一行开始）
    data_rows = rows[header_row_idx+1:]
    records = []  # 每个元素：{group, time, name, others}
    for r_idx, row_vals in enumerate(data_rows):
        if not any(row_vals):  # 跳过全空行
            continue
        for g_idx, g in enumerate(groups):
            time_val = row_vals[g['time_col']] if g['time_col'] < len(row_vals) else None
            name_val = row_vals[g['name_col']] if g['name_col'] < len(row_vals) else None
            if time_val is None or name_val is None:
                # 缺少关键信息，跳过
                continue
            # 时间转为浮点
            try:
                t = float(time_val)
            except:
                continue
            name = str(name_val).strip()
            # 其他列值
            others = [row_vals[c] if c < len(row_vals) else None for c in g['other_cols']]
            records.append({
                'group': g_idx,
                'time': t,
                'name': name,
                'others': others
            })

    # 5. 对齐（按时间排序，按名称和时间差匹配）
    records.sort(key=lambda x: x['time'])
    used = [False] * len(records)
    merged_rows = []  # 每个元素：{'name': name, 'group_data': {g_idx: (time, others)} }

    # 使用 tqdm 显示合并进度
    iter_records = tqdm(records, desc="合并对齐记录", disable=not show_progress, unit="条")
    for i, rec in enumerate(iter_records):
        if used[i]:
            continue
        # 创建新行
        name = rec['name']
        group_data = {rec['group']: (rec['time'], rec['others'])}
        used[i] = True
        # 向后查找可以合并的记录
        for j in range(i+1, len(records)):
            if used[j]:
                continue
            rj = records[j]
            if rj['name'] == name and abs(rj['time'] - rec['time']) <= dev and rj['group'] != rec['group']:
                # 检查该组是否已被占用（一个组只能有一个数据）
                if rj['group'] in group_data:
                    continue
                group_data[rj['group']] = (rj['time'], rj['others'])
                used[j] = True
        merged_rows.append({
            'name': name,
            'group_data': group_data
        })

    # 6. 构建新表头
    # 第一列标题：取第一个名称列的原始标题，或 "Compound"
    first_name_header = header_row_vals[groups[0]['name_col']] if groups else 'Compound'
    name_header = str(first_name_header).strip() if first_name_header else 'Compound'
    new_headers = [name_header]
    # 对于每个组，按顺序取除名称列外的列标题
    for g_idx, g in enumerate(groups):
        # 本组的其他列（包括时间和其他）
        other_cols = [g['time_col']] + g['other_cols']   # 时间列放在最前
        for col_idx in other_cols:
            title = header_row_vals[col_idx] if col_idx < len(header_row_vals) else ''
            title = str(title).strip() if title else ''
            if not title:
                title = f'Col_{col_idx+1}'
            # 加组序号避免重复
            suffix = f'_{g_idx+1}'
            new_title = title + suffix
            new_headers.append(new_title)

    # 7. 构建数据行（每行对应 merged_rows 中的一个）
    data_lines = []
    for mr in merged_rows:
        line = [mr['name']]
        for g_idx, g in enumerate(groups):
            if g_idx in mr['group_data']:
                time_val, others = mr['group_data'][g_idx]
                # 顺序：时间列 + other_cols
                line.append(time_val)
                line.extend(others)
            else:
                # 无数据，填空
                line.append(None)
                line.extend([None] * len(g['other_cols']))
        data_lines.append(line)

    return above_rows, new_headers, data_lines

def out():
    option = json.loads(sys.argv[1])
    file = option["File"]
    dev = float(option.get("Para1", 0.1))
    count = option.get("Para2", 0)
    count_param = None if count == "" else int(count)

    wb = openpyxl.load_workbook(file)
    out_wb = openpyxl.Workbook()
    # 删除默认的 Sheet
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    sheet_names = wb.sheetnames
    # 总进度：处理工作表
    for sheet_name in tqdm(sheet_names, desc="处理工作表", unit="个"):
        ws = wb[sheet_name]
        above_rows, new_headers, data_lines = process_sheet(ws, dev, count_param, show_progress=True)

        # 创建新 sheet
        out_ws = out_wb.create_sheet(title=sheet_name)

        # 写入上方行
        for r_idx, row_vals in enumerate(above_rows):
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=r_idx+1, column=c_idx+1, value=val)

        # 写入新表头（紧接着上方行之后）
        header_row_num = len(above_rows) + 1
        for c_idx, val in enumerate(new_headers):
            out_ws.cell(row=header_row_num, column=c_idx+1, value=val)

        # 写入数据行
        for r_idx, row_vals in enumerate(data_lines):
            row_num = header_row_num + 1 + r_idx
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=row_num, column=c_idx+1, value=val)

    # 保存
    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}.xlsx")
    out_wb.save(out_path)
    print(f"处理完成，文件保存至：{out_path}")

if __name__ == "__main__":
    out()