import os
import openpyxl
import time
import sys
import json
import math
import re
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
from PyQt5.QtWidgets import QApplication, QMessageBox

# ================= 公共识别函数 =================
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]

TIME_KEYWORDS = ('保留时间', '时间', 'time')
NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')
CAS_KEYWORDS = ('cas', 'cas号', 'cas number')

def is_time_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in TIME_KEYWORDS)

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in NAME_KEYWORDS)

def is_cas_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in CAS_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    """返回同时包含时间列和名称列的行索引（0-based）。"""
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r+1]]
        has_time = any(is_time_header(v) for v in row_vals)
        has_name = any(is_name_header(v) for v in row_vals)
        if has_time and has_name:
            return r
    return 0

# ================= 正构烷烃参考文件加载 =================
def load_alkanes(file_path):
    """读取正构烷烃参考文件，返回 [(carbon_num, rt), ...] 按碳数升序排列。"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    alkanes = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if len(row) < 2:
            continue
        carbon_val, rt_val = row[0], row[1]
        if carbon_val is None or rt_val is None:
            continue
        try:
            carbon_num = int(carbon_val)
            rt = float(rt_val)
        except (ValueError, TypeError):
            continue
        if carbon_num > 0 and rt > 0:
            alkanes.append((carbon_num, rt))
    wb.close()
    if len(alkanes) < 2:
        raise ValueError(f"正构烷烃参考文件中有效数据不足（至少需要2个），当前读取到 {len(alkanes)} 个。")
    alkanes.sort(key=lambda x: x[0])
    return alkanes

# ================= 保留指数计算 =================
def calc_ri_kovats(t_x, alkanes):
    """Kovats对数插值公式计算保留指数。"""
    if t_x <= 0:
        return None
    # 等于某个烷烃
    for carbon_num, rt in alkanes:
        if abs(t_x - rt) < 1e-10:
            return carbon_num * 100
    # 低于范围
    if t_x < alkanes[0][1]:
        return "低于范围"
    # 超出范围
    if t_x > alkanes[-1][1]:
        return "超出范围"
    # 查找包围的两个烷烃
    for i in range(len(alkanes) - 1):
        z, t_z = alkanes[i]
        z1, t_z1 = alkanes[i + 1]
        if t_z <= t_x < t_z1:
            if t_z <= 0 or t_z1 <= 0:
                return None
            ri = 100 * (z + (math.log10(t_x) - math.log10(t_z)) / (math.log10(t_z1) - math.log10(t_z)))
            return round(ri, 1)
    return None

def calc_ri_linear(t_x, alkanes):
    """线性插值公式计算保留指数。"""
    if t_x <= 0:
        return None
    for carbon_num, rt in alkanes:
        if abs(t_x - rt) < 1e-10:
            return carbon_num * 100
    if t_x < alkanes[0][1]:
        return "低于范围"
    if t_x > alkanes[-1][1]:
        return "超出范围"
    for i in range(len(alkanes) - 1):
        z, t_z = alkanes[i]
        z1, t_z1 = alkanes[i + 1]
        if t_z <= t_x < t_z1:
            if t_z1 == t_z:
                return None
            ri = 100 * (z + (t_x - t_z) / (t_z1 - t_z))
            return round(ri, 1)
    return None

# ================= NIST RI 查询 =================
def get_nist_ri(cas, calc_ri):
    """根据CAS号查询NIST数据库中最接近calc_ri的保留指数。"""
    if cas is None or str(cas).strip() == "":
        return "\\"
    clean_cas = str(cas).strip().replace('-', '')
    url = f"https://webbook.nist.gov/cgi/cbook.cgi?ID=C{clean_cas}&Units=SI&Mask=2000"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    try:
        response = requests.get(url, headers=headers, timeout=15)
        if response.status_code != 200:
            return "\\"
        soup = BeautifulSoup(response.text, 'html.parser')
        target_title = "Normal alkane RI, non-polar column, custom temperature program"
        target_table = soup.find('table', attrs={'aria-label': target_title})
        if not target_table:
            text_matches = soup.find_all(string=re.compile(re.escape(target_title)))
            for match in text_matches:
                parent = match.parent
                next_table = parent.find_next('table')
                if next_table:
                    target_table = next_table
                    break
        if not target_table:
            return "\\"
        header_row = target_table.find('tr')
        headers_text = [th.get_text(strip=True) for th in header_row.find_all(['th', 'td'])]
        try:
            ri_col_index = headers_text.index('I')
        except ValueError:
            return "\\"
        ri_values = []
        rows = target_table.find_all('tr')[1:]
        for row in rows:
            cols = row.find_all(['td', 'th'])
            if len(cols) > ri_col_index:
                val_text = cols[ri_col_index].get_text(strip=True)
                val_clean = re.sub(r'[^\d\.]', '', val_text)
                try:
                    if val_clean:
                        ri_values.append(float(val_clean))
                except ValueError:
                    continue
        if not ri_values:
            return "\\"
        closest_ri = min(ri_values, key=lambda x: abs(x - float(calc_ri)))
        if closest_ri.is_integer():
            return int(closest_ri)
        else:
            return closest_ri
    except Exception as e:
        print(f"  查询CAS {cas} 时出错: {e}")
        return "\\"

# ================= 工作表处理 =================
def process_sheet(ws, alkanes, formula, agg, cas_col_param, name_col_param, show_progress=True):
    """
    处理单个工作表，计算保留指数。
    返回 (above_rows, new_header, new_data_rows, ri_col_indices)
    """
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        rows.append(row_vals)

    # 检测表头行
    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx + 1:]

    # 确定名称列索引
    if name_col_param:
        name_col = HEAD.index(name_col_param.upper())
    else:
        name_col = None
        for i, val in enumerate(header_row_vals):
            if is_name_header(val):
                name_col = i
                break
        if name_col is None:
            raise ValueError(f"工作表 '{ws.title}' 中未找到名称列，请通过参数指定。")

    # 确定CAS列索引（可选，用于NIST查询）
    cas_col = None
    if cas_col_param:
        cas_col = HEAD.index(cas_col_param.upper())
    else:
        for i, val in enumerate(header_row_vals):
            if is_cas_header(val):
                cas_col = i
                break

    # 找到所有保留时间列
    time_cols = [i for i, val in enumerate(header_row_vals) if is_time_header(val)]
    if not time_cols:
        raise ValueError(f"工作表 '{ws.title}' 中未找到保留时间列（表头需包含'保留时间'/'时间'/'time'）。")

    # 确定插入RI列的位置（名称列右侧）
    ri_insert_pos = name_col + 1

    # 构建新表头
    new_header = list(header_row_vals)
    if formula in ("L", "K"):
        ri_col_name = "保留指数(线性)" if formula == "L" else "保留指数(Kovats)"
        new_header.insert(ri_insert_pos, ri_col_name)
        ri_col_count = 1
    else:  # B - both
        new_header.insert(ri_insert_pos, "保留指数(Kovats)")
        new_header.insert(ri_insert_pos + 1, "保留指数(线性)")
        ri_col_count = 2

    # 构建新数据行
    new_data_rows = []
    for row in tqdm(data_rows, desc=f"处理 {ws.title}", disable=not show_progress, unit="行"):
        # 收集该行所有有效保留时间
        rts = []
        for tc in time_cols:
            if tc < len(row) and row[tc] is not None:
                try:
                    rts.append(float(row[tc]))
                except (ValueError, TypeError):
                    pass

        # 聚合保留时间
        if not rts:
            agg_rt = None
        elif agg == "A":
            agg_rt = sum(rts) / len(rts)
        elif agg == "MIN":
            agg_rt = min(rts)
        elif agg == "MAX":
            agg_rt = max(rts)
        else:
            agg_rt = sum(rts) / len(rts)

        # 计算RI
        if agg_rt is None:
            ri_kovats = ""
            ri_linear = ""
        else:
            if formula in ("K", "B"):
                ri_kovats = calc_ri_kovats(agg_rt, alkanes)
            else:
                ri_kovats = None
            if formula in ("L", "B"):
                ri_linear = calc_ri_linear(agg_rt, alkanes)
            else:
                ri_linear = None

        # 构建新行
        new_row = list(row)
        if formula == "L":
            new_row.insert(ri_insert_pos, ri_linear)
        elif formula == "K":
            new_row.insert(ri_insert_pos, ri_kovats)
        else:  # B
            new_row.insert(ri_insert_pos, ri_kovats)
            new_row.insert(ri_insert_pos + 1, ri_linear)
        new_data_rows.append(new_row)

    return above_rows, new_header, new_data_rows, cas_col, ri_insert_pos, ri_col_count

# ================= 输出写入 =================
def write_output(wb, sheets_results, suffix, show_progress=True):
    """将处理结果写入新工作簿并保存。"""
    out_wb = openpyxl.Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for sheet_name, (above_rows, new_header, new_data_rows) in tqdm(
        sheets_results.items(), desc="写入工作表", disable=not show_progress, unit="个"
    ):
        out_ws = out_wb.create_sheet(title=sheet_name)

        # 写入上方行
        for r_idx, row_vals in enumerate(above_rows):
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)

        # 写入新表头
        header_row_num = len(above_rows) + 1
        for c_idx, val in enumerate(new_header):
            if val is not None:
                out_ws.cell(row=header_row_num, column=c_idx + 1, value=val)

        # 写入数据行
        for r_idx, row_vals in enumerate(new_data_rows):
            row_num = header_row_num + 1 + r_idx
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=row_num, column=c_idx + 1, value=val)

    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}{suffix}.xlsx")
    out_wb.save(out_path)
    return out_path

# ================= NIST 查询后处理 =================
def add_nist_ri_to_file(file_path, cas_col_map, ri_insert_pos_map, ri_col_count_map):
    """读取已保存的文件，查询NIST RI并添加新列。"""
    wb = openpyxl.load_workbook(file_path)
    out_wb = openpyxl.Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column

        cas_col = cas_col_map.get(sheet_name)
        ri_insert_pos = ri_insert_pos_map.get(sheet_name, 0)
        ri_col_count = ri_col_count_map.get(sheet_name, 1)

        if cas_col is None:
            # 没有CAS列，直接复制
            out_ws = out_wb.create_sheet(title=sheet_name)
            for r in range(1, max_row + 1):
                for c in range(1, max_col + 1):
                    val = ws.cell(row=r, column=c).value
                    if val is not None:
                        out_ws.cell(row=r, column=c, value=val)
            continue

        # 找到RI列中可用于对比的值（取第一个RI列）
        ri_col_for_compare = ri_insert_pos  # 第一个RI列的0-based索引

        # 读取所有行
        rows = []
        for r in range(1, max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            rows.append(row_vals)

        # 检测表头行
        header_idx = detect_header_row(ws)

        # NIST RI插入位置：在最后一个RI列之后
        nist_insert_pos = ri_insert_pos + ri_col_count  # 0-based index

        # 构建新表头
        new_header = list(rows[header_idx])
        new_header.insert(nist_insert_pos, "NIST RI")

        # 构建新数据行
        new_data_rows = []
        data_rows = rows[header_idx + 1:]
        for r_idx, row in enumerate(tqdm(data_rows, desc=f"NIST查询 {sheet_name}", unit="行")):
            cas_val = row[cas_col] if cas_col < len(row) else None
            ri_val = row[ri_col_for_compare] if ri_col_for_compare < len(row) else None

            # 获取用于对比的RI值
            calc_ri_for_compare = None
            if ri_val is not None:
                try:
                    calc_ri_for_compare = float(ri_val)
                except (ValueError, TypeError):
                    pass

            if cas_val and calc_ri_for_compare is not None and calc_ri_for_compare > 0:
                nist_ri = get_nist_ri(cas_val, calc_ri_for_compare)
                print(f"  [{r_idx+1}] CAS: {cas_val} | 计算RI: {calc_ri_for_compare} | NIST RI: {nist_ri}")
                time.sleep(1)
            else:
                nist_ri = ""

            new_row = list(row)
            new_row.insert(nist_insert_pos, nist_ri)
            new_data_rows.append(new_row)

        # 写入
        out_ws = out_wb.create_sheet(title=sheet_name)
        above_rows = rows[:header_idx]
        for r_idx, row_vals in enumerate(above_rows):
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)

        header_row_num = len(above_rows) + 1
        for c_idx, val in enumerate(new_header):
            if val is not None:
                out_ws.cell(row=header_row_num, column=c_idx + 1, value=val)

        for r_idx, row_vals in enumerate(new_data_rows):
            row_num = header_row_num + 1 + r_idx
            for c_idx, val in enumerate(row_vals):
                if val is not None:
                    out_ws.cell(row=row_num, column=c_idx + 1, value=val)

    # 保存
    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}_with_RI_NIST.xlsx")
    out_wb.save(out_path)
    wb.close()
    return out_path

# ================= 主程序 =================
def main():
    option = json.loads(sys.argv[1])
    file = option["File"]
    para1 = option.get("Para1", "").strip()  # 参考文件路径
    para2 = option.get("Para2", "").strip()  # 公式,聚合
    para3 = option.get("Para3", "").strip()  # CAS列,名称列

    # 解析Para1：参考文件路径
    if not para1:
        print("错误：请提供正构烷烃参考文件路径（Para1）。")
        return
    if not os.path.exists(para1):
        print(f"错误：参考文件不存在: {para1}")
        return

    # 解析Para2：公式和聚合方式
    if para2:
        parts = [p.strip().upper() for p in para2.split(",")]
        formula = parts[0] if len(parts) > 0 and parts[0] in ("L", "K", "B") else "L"
        agg = parts[1] if len(parts) > 1 and parts[1] in ("A", "MIN", "MAX") else "A"
    else:
        formula, agg = "L", "A"

    # 解析Para3：CAS列和名称列
    cas_col_param = None
    name_col_param = None
    if para3:
        parts = [p.strip() for p in para3.split(",")]
        if len(parts) >= 1 and parts[0]:
            cas_col_param = parts[0].upper()
        if len(parts) >= 2 and parts[1]:
            name_col_param = parts[1].upper()

    # 加载正构烷烃参考数据
    try:
        alkanes = load_alkanes(para1)
        print(f"已加载 {len(alkanes)} 个正构烷烃数据（碳数 {alkanes[0][0]}-{alkanes[-1][0]}）")
    except Exception as e:
        print(f"错误：加载参考文件失败 - {e}")
        return

    # 读取输入文件
    if not os.path.exists(file):
        print(f"错误：输入文件不存在: {file}")
        return

    wb = openpyxl.load_workbook(file)
    sheets_results = {}
    cas_col_map = {}
    ri_insert_pos_map = {}
    ri_col_count_map = {}

    # 处理每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            above_rows, new_header, new_data_rows, cas_col, ri_insert_pos, ri_col_count = process_sheet(
                ws, alkanes, formula, agg, cas_col_param, name_col_param
            )
            sheets_results[sheet_name] = (above_rows, new_header, new_data_rows)
            cas_col_map[sheet_name] = cas_col
            ri_insert_pos_map[sheet_name] = ri_insert_pos
            ri_col_count_map[sheet_name] = ri_col_count
        except ValueError as e:
            print(f"工作表 '{sheet_name}' 处理失败: {e}")
            continue

    wb.close()

    if not sheets_results:
        print("错误：没有成功处理任何工作表。")
        return

    # 写入第一阶段结果
    out_path = write_output(wb=None, sheets_results=sheets_results, suffix="_with_RI")
    print(f"\nRI计算完成！文件保存至：{out_path}")

    # 询问是否查询NIST
    has_cas = any(v is not None for v in cas_col_map.values())
    if has_cas:
        app = QApplication.instance() or QApplication(sys.argv)
        reply = QMessageBox.question(
            None, "NIST RI 查询",
            "RI计算完成！是否查询NIST数据库进行RI对比？\n\n（查询过程较慢，每条约1秒）",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        if reply == QMessageBox.Yes:
            print("\n开始NIST查询（每条约1秒，请耐心等待）...")
            nist_path = add_nist_ri_to_file(out_path, cas_col_map, ri_insert_pos_map, ri_col_count_map)
            print(f"\nNIST查询完成！最终文件保存至：{nist_path}")
    else:
        print("\n（未检测到CAS号列，跳过NIST查询。如需NIST查询，请通过Para3指定CAS列位置。）")

if __name__ == "__main__":
    main()
