import os
import openpyxl
import time
import sys
import json

head = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z",
"AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ",
"BA","BB","BC","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BM","BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ",
"CA","CB","CC","CD","CE","CF","CG","CH","CI","CJ","CK","CL","CM","CN","CO","CP","CQ","CR","CS","CT","CU","CV","CW","CX","CY","CZ",
"DA","DB","DC","DD","DE","DF","DG","DH","DI","DJ","DK","DL","DM","DN","DO","DP","DQ","DR","DS","DT","DU","DV","DW","DX","DY","DZ",
"EA","EB","EC","ED","EE","EF","EG","EH","EI","EJ","EK","EL","EM","EN","EO","EP","EQ","ER","ES","ET","EU","EV","EW","EX","EY","EZ"]

#数据处理函数
def process(workbook,sheetname,dev,count):
    sheet = workbook[f'{sheetname}']
    state = {}
    order = {}
    sequence = {}
    listhead = []
    #读取数据
    for i in head:
        cell = sheet[f"{i}"]
        state[f"{i}"] = []
        for j in cell:
            state[f"{i}"].append(j.value)
    for i in head:
        listhead.append(state[f"{i}"][0])
        del state[f"{i}"][0]

    #格式化数据
    for k in range(0,156,count):
        i = head[k]
        for j in range(len(state[f"{i}"])):
            order[f"{i}.{j}"] = state[f"{i}"][j]
    for k in range(0,156,count):
        i = head[k]
        for j in range(len(state[f"{i}"])):
            if order[f"{i}.{j}"] == None:
                del order[f"{i}.{j}"]

    #排序数据
    ordered=list(order.items())       # 得到列表
    ordered.sort(key=lambda x:x[1],reverse=False)  # 按列表中，每一个元组的第二个元素从小到大排序。x代表从L中遍历出的一个元组

    #将数据分配到各行
    l = 1
    marker = []
    for i in ordered:
        h,n = i[0].split(".")
        hh = head[head.index(h) + 1]
        n = int(n)
        name = state[f"{hh}"][n]
        if f"{h}.{n}" in marker:
            continue
        sequence[f"{l}"] = [f"{h}.{n}"]
        marker.append(f"{h}.{n}")
        lst = head[0::count]
        for j in lst:
            if j == h:
                continue
            jj = head[head.index(j) + 1]
            for k in range(len(state[f"{j}"])):
                num = state[f"{j}"][k]
                if num == None:
                    continue
                if state[f"{jj}"][k] == name and abs(num - i[1]) <= dev:
                    sequence[f"{l}"].append(f"{j}.{k}")
                    marker.append(f"{j}.{k}")
        l = l + 1

    #格式化输出
    output = {}
    for i in range(1,l):
        mark = sequence[f"{i}"]
        output[f"{i}"] = [None] * 156
        for j in mark:
            column,line = j.split(".")
            line = int(line)
            st = head.index(column)
            en = st + count
            for k in range(st,en):
                output[f"{i}"][k] = state[f"{head[k]}"][line]
    output["0"] = listhead
    return output

def out():
    option = json.loads(sys.argv[1])
    file = option["File"]
    print(f"目标路径：{file}")
    dev = float(option["Para1"])
    count = int(option["Para2"])
    workbook = openpyxl.load_workbook(f'{file}')
    sheetlist = workbook.sheetnames
    
    #处理数据并保存至新文件
    try:
        outbook = openpyxl.Workbook()
        for sheetname in sheetlist:
            outbook.create_sheet(f'{sheetname}')
            outsheet = outbook[f'{sheetname}']
            output = process(workbook,sheetname,dev,count)
            for i in range(len(output)):
                line = output[f"{i}"]
                j = i + 1
                for k in range(len(line)):
                    outsheet[f"{head[k]}{j}"] = line[k]
        path = "./output/"
        if not os.path.exists(path):
            os.mkdir(path)
        filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
        outbook.save(f'{path}{filename}.xlsx')
        print(f'已完成处理，文件名为{path}{filename}.xlsx')
    except Exception as r:
        print('错误:',f'{r}')

if __name__ == "__main__":
    out()