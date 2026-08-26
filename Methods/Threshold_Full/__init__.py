# -*- coding: utf-8 -*-
"""
Threshold_Full 模块
基于 Threshold_full.json 完整数据库的高级阈值查询模块。
支持三种模式：default（默认最新值）、custom（自定义筛选）、manual（手动选择）

参数说明：
  Para1 - （可选）CAS号列位置（列字母，如 D）
  Para2 - 模式：d(默认) / c(自定义) / m(手动)
  Para3 - 无
"""

import os
import sys
import json
import re
import time
import openpyxl
from tqdm import tqdm
from PyQt5.QtWidgets import (QApplication, QDialog, QVBoxLayout, QHBoxLayout,
                              QLabel, QComboBox, QCheckBox, QPushButton,
                              QTableWidget, QTableWidgetItem, QHeaderView,
                              QAbstractItemView, QMessageBox)
from PyQt5.QtCore import Qt

# ================= 列字母工具 =================
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]

# ================= 表头识别关键词 =================
TIME_KEYWORDS = ('保留时间', '时间', 'time')
NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')
CAS_KEYWORDS = ('cas', 'cas号', 'cas number')

def is_time_header(val):
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and any(kw in s.lower() for kw in TIME_KEYWORDS)

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and any(kw in s.lower() for kw in NAME_KEYWORDS)

def is_cas_header(val):
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and any(kw in s.lower() for kw in CAS_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r + 1]]
        has_time = any(is_time_header(v) for v in row_vals)
        has_name = any(is_name_header(v) for v in row_vals)
        if has_time and has_name:
            return r
    return 0

# ================= 数据库加载 =================
def load_full_thresholds(json_path="./Database/Threshold_full.json"):
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"未找到阈值数据库文件: {json_path}")
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    cas_index = {}
    name_index = {}
    for compound in data:
        cas = (compound.get('cas_number') or '').strip()
        if cas:
            cas_index[cas] = compound
        en = (compound.get('english_name') or '').strip().lower()
        cn = (compound.get('chinese_name') or '').strip()
        if en:
            name_index[en] = compound
        if cn:
            name_index[cn] = compound
    return data, cas_index, name_index

# ================= 范围值解析 =================
def parse_threshold_value(val_str):
    """解析阈值字符串，返回数值列表。支持范围值如 '0.001～0.002'"""
    if not val_str:
        return []
    s = str(val_str).strip()
    s = re.sub(r'^[<>≤≥]', '', s)
    range_match = re.match(r'([\d.]+)\s*[～~\-]\s*([\d.]+)', s)
    if range_match:
        try:
            return [float(range_match.group(1)), float(range_match.group(2))]
        except ValueError:
            return []
    try:
        return [float(s)]
    except ValueError:
        return []

def get_max_value(thresholds):
    """获取阈值列表中的最大值"""
    max_val = None
    for t in thresholds:
        vals = parse_threshold_value(t.get('value', ''))
        if vals:
            v = max(vals)
            if max_val is None or v > max_val:
                max_val = v
    return max_val

def get_min_value(thresholds):
    """获取阈值列表中的最小值"""
    min_val = None
    for t in thresholds:
        vals = parse_threshold_value(t.get('value', ''))
        if vals:
            v = min(vals)
            if min_val is None or v < min_val:
                min_val = v
    return min_val

def get_latest_threshold(thresholds):
    """获取最新年份的阈值"""
    if not thresholds:
        return None
    valid = [t for t in thresholds if t.get('year')]
    if not valid:
        return thresholds[0]
    max_year = max(t.get('year', 0) for t in valid)
    latest = [t for t in valid if t.get('year') == max_year]
    return latest[0] if latest else thresholds[0]

def filter_by_type(thresholds, type_filter):
    """按类型筛选阈值"""
    if not type_filter or type_filter == '全部':
        return thresholds
    result = []
    for t in thresholds:
        t_type = (t.get('type') or '').strip().lower()
        if not t_type or t_type == type_filter.lower():
            result.append(t)
    return result

# ================= PyQt5 窗口 =================
class CustomModeWindow(QDialog):
    """自定义模式设置窗口"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("自定义模式 - 筛选条件")
        self.setFixedSize(350, 250)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()

        row1 = QHBoxLayout()
        row1.addWidget(QLabel("匹配方式:"))
        self.match_combo = QComboBox()
        self.match_combo.addItems(["CAS 号", "名称"])
        row1.addWidget(self.match_combo)
        layout.addLayout(row1)

        row2 = QHBoxLayout()
        row2.addWidget(QLabel("阈值选择:"))
        self.value_combo = QComboBox()
        self.value_combo.addItems(["最新值", "最大值", "最小值"])
        row2.addWidget(self.value_combo)
        layout.addLayout(row2)

        row3 = QHBoxLayout()
        row3.addWidget(QLabel("类型:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems(["全部", "detection", "recognition"])
        row3.addWidget(self.type_combo)
        layout.addLayout(row3)

        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确认")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def get_settings(self):
        return {
            'match_mode': 'cas' if self.match_combo.currentIndex() == 0 else 'name',
            'value_mode': ['latest', 'max', 'min'][self.value_combo.currentIndex()],
            'type_filter': self.type_combo.currentText()
        }

class ManualModeSettingsWindow(QDialog):
    """手动模式设置窗口"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("手动模式 - 匹配方式")
        self.setFixedSize(300, 180)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout()
        layout.addWidget(QLabel("请选择匹配方式（可多选）:"))

        self.cas_check = QCheckBox("CAS 号")
        self.cas_check.setChecked(True)
        layout.addWidget(self.cas_check)

        self.name_check = QCheckBox("名称")
        self.name_check.setChecked(True)
        layout.addWidget(self.name_check)

        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确认")
        ok_btn.clicked.connect(self.validate_and_accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def validate_and_accept(self):
        if not self.cas_check.isChecked() and not self.name_check.isChecked():
            QMessageBox.warning(self, "提示", "请至少选择一种匹配方式")
            return
        self.accept()

    def get_settings(self):
        return {
            'use_cas': self.cas_check.isChecked(),
            'use_name': self.name_check.isChecked()
        }

class ManualModeDialog(QDialog):
    """手动模式逐行选择对话框"""
    def __init__(self, compound_name, cas_number, thresholds, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"选择阈值 - {compound_name}")
        self.setMinimumSize(600, 400)
        self.thresholds = thresholds
        self.selected_indices = []
        self.setup_ui(compound_name, cas_number)

    def setup_ui(self, compound_name, cas_number):
        layout = QVBoxLayout()

        info = f"化合物: {compound_name}"
        if cas_number:
            info += f"  |  CAS: {cas_number}"
        layout.addWidget(QLabel(info))

        self.table = QTableWidget(len(self.thresholds), 5)
        self.table.setHorizontalHeaderLabels(["值", "单位", "研究者", "年份", "类型"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.MultiSelection)

        for i, t in enumerate(self.thresholds):
            self.table.setItem(i, 0, QTableWidgetItem(str(t.get('value', ''))))
            self.table.setItem(i, 1, QTableWidgetItem(str(t.get('unit', ''))))
            self.table.setItem(i, 2, QTableWidgetItem(str(t.get('researcher', ''))))
            self.table.setItem(i, 3, QTableWidgetItem(str(t.get('year', ''))))
            self.table.setItem(i, 4, QTableWidgetItem(str(t.get('type', ''))))

        layout.addWidget(self.table)

        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确认选择")
        ok_btn.clicked.connect(self.accept_selection)
        skip_btn = QPushButton("跳过")
        skip_btn.clicked.connect(self.skip)
        cancel_btn = QPushButton("取消处理")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(skip_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def accept_selection(self):
        selected = self.table.selectionModel().selectedRows()
        self.selected_indices = [idx.row() for idx in selected]
        self.done(QDialog.Accepted)

    def skip(self):
        self.selected_indices = []
        self.done(QDialog.Accepted)

    def get_selected_values(self):
        if not self.selected_indices:
            return "Skipped"
        values = []
        for idx in self.selected_indices:
            val = self.thresholds[idx].get('value', '')
            if val:
                values.append(str(val))
        return '; '.join(values) if values else "Skipped"

# ================= 核心处理逻辑 =================
def find_columns(header_row_vals):
    """自动识别 CAS 列和名称列"""
    cas_col = None
    name_col = None
    for i, val in enumerate(header_row_vals):
        if cas_col is None and is_cas_header(val):
            cas_col = i
        if name_col is None and is_name_header(val):
            name_col = i
    return cas_col, name_col

def lookup_by_cas(cas_val, cas_index):
    if not cas_val:
        return None
    cas_key = str(cas_val).strip()
    return cas_index.get(cas_key)

def lookup_by_name(name_val, name_index):
    if not name_val:
        return None
    name_key = str(name_val).strip().lower()
    return name_index.get(name_key)

def process_default_mode(ws, cas_col_param, cas_index, name_index):
    """默认模式：自动匹配最新阈值"""
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        rows.append(row_vals)

    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx + 1:]

    auto_cas_col, auto_name_col = find_columns(header_row_vals)
    if cas_col_param:
        cas_col = HEAD.index(cas_col_param.upper())
    else:
        cas_col = auto_cas_col
        if cas_col is None:
            raise ValueError("未找到 CAS 列，请通过参数1指定列字母。")

    name_col = auto_name_col

    new_header = list(header_row_vals)
    insert_pos = cas_col + 1
    new_header.insert(insert_pos, "嗅觉阈值(mg/L)")

    new_data_rows = []
    stats = {'matched': 0, 'not_found': 0}

    for row in tqdm(data_rows, desc="处理数据", unit="行"):
        cas_val = row[cas_col] if cas_col < len(row) else None
        name_val = row[name_col] if name_col is not None and name_col < len(row) else None

        if cas_val is None and name_val is None:
            new_row = list(row)
            new_row.insert(insert_pos, "")
            new_data_rows.append(new_row)
            continue

        compound = lookup_by_cas(cas_val, cas_index)
        if compound is None and name_val:
            compound = lookup_by_name(name_val, name_index)

        if compound is None:
            new_row = list(row)
            new_row.insert(insert_pos, "Not Found")
            new_data_rows.append(new_row)
            stats['not_found'] += 1
            continue

        thresholds = compound.get('thresholds', [])
        if not thresholds:
            new_row = list(row)
            new_row.insert(insert_pos, "No Data")
            new_data_rows.append(new_row)
            continue

        latest = get_latest_threshold(thresholds)
        new_row = list(row)
        new_row.insert(insert_pos, str(latest.get('value', '')) if latest else "No Data")
        new_data_rows.append(new_row)
        stats['matched'] += 1

    return above_rows, new_header, new_data_rows, stats

def process_custom_mode(ws, cas_col_param, cas_index, name_index, settings):
    """自定义模式：根据用户设置筛选"""
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        rows.append(row_vals)

    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx + 1:]

    if cas_col_param:
        lookup_col = HEAD.index(cas_col_param.upper())
        lookup_mode = 'cas'
    else:
        cas_col, name_col = find_columns(header_row_vals)
        if settings['match_mode'] == 'cas':
            lookup_col = cas_col
            lookup_mode = 'cas'
        else:
            lookup_col = name_col
            lookup_mode = 'name'
        if lookup_col is None:
            raise ValueError(f"未找到{'CAS' if lookup_mode == 'cas' else '名称'}列，请通过参数1指定。")

    new_header = list(header_row_vals)
    insert_pos = lookup_col + 1
    new_header.insert(insert_pos, "嗅觉阈值(mg/L)")

    new_data_rows = []
    stats = {'matched': 0, 'not_found': 0}

    for row in tqdm(data_rows, desc="处理数据", unit="行"):
        lookup_val = row[lookup_col] if lookup_col < len(row) else None

        if lookup_val is None:
            new_row = list(row)
            new_row.insert(insert_pos, "")
            new_data_rows.append(new_row)
            continue

        if lookup_mode == 'cas':
            compound = lookup_by_cas(lookup_val, cas_index)
        else:
            compound = lookup_by_name(lookup_val, name_index)

        if compound is None:
            new_row = list(row)
            new_row.insert(insert_pos, "Not Found")
            new_data_rows.append(new_row)
            stats['not_found'] += 1
            continue

        thresholds = compound.get('thresholds', [])
        if not thresholds:
            new_row = list(row)
            new_row.insert(insert_pos, "No Data")
            new_data_rows.append(new_row)
            continue

        filtered = filter_by_type(thresholds, settings.get('type_filter'))

        value_mode = settings.get('value_mode', 'latest')
        if value_mode == 'latest':
            result = get_latest_threshold(filtered)
            value = str(result.get('value', '')) if result else "No Match"
        elif value_mode == 'max':
            value = str(get_max_value(filtered)) if filtered else "No Match"
        else:
            value = str(get_min_value(filtered)) if filtered else "No Match"

        new_row = list(row)
        new_row.insert(insert_pos, value)
        new_data_rows.append(new_row)
        stats['matched'] += 1

    return above_rows, new_header, new_data_rows, stats

def process_manual_mode(ws, cas_col_param, cas_index, name_index, settings):
    """手动模式：用户逐行选择"""
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        rows.append(row_vals)

    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx + 1:]

    auto_cas_col, auto_name_col = find_columns(header_row_vals)
    if cas_col_param:
        cas_col = HEAD.index(cas_col_param.upper())
    else:
        cas_col = auto_cas_col
    name_col = auto_name_col

    insert_col = cas_col if cas_col is not None else (name_col if name_col is not None else 0)
    insert_pos = insert_col + 1

    new_header = list(header_row_vals)
    new_header.insert(insert_pos, "嗅觉阈值(mg/L)")

    new_data_rows = []
    stats = {'matched': 0, 'not_found': 0, 'skipped': 0}
    cancelled = False

    for row in tqdm(data_rows, desc="处理数据", unit="行"):
        if cancelled:
            new_row = list(row)
            new_row.insert(insert_pos, "")
            new_data_rows.append(new_row)
            continue

        cas_val = row[cas_col] if cas_col is not None and cas_col < len(row) else None
        name_val = row[name_col] if name_col is not None and name_col < len(row) else None

        if cas_val is None and name_val is None:
            new_row = list(row)
            new_row.insert(insert_pos, "")
            new_data_rows.append(new_row)
            continue

        compound = None
        use_cas = settings.get('use_cas', True)
        use_name = settings.get('use_name', True)

        if use_cas and cas_val:
            compound = lookup_by_cas(cas_val, cas_index)
        if compound is None and use_name and name_val:
            compound = lookup_by_name(name_val, name_index)

        if compound is None:
            new_row = list(row)
            new_row.insert(insert_pos, "Not Found")
            new_data_rows.append(new_row)
            stats['not_found'] += 1
            continue

        thresholds = compound.get('thresholds', [])
        if not thresholds:
            new_row = list(row)
            new_row.insert(insert_pos, "No Data")
            new_data_rows.append(new_row)
            continue

        compound_name = compound.get('chinese_name') or compound.get('english_name', '')
        cas_number = compound.get('cas_number', '')
        dialog = ManualModeDialog(compound_name, cas_number, thresholds)
        result = dialog.exec_()

        if result == QDialog.Rejected:
            cancelled = True
            new_row = list(row)
            new_row.insert(insert_pos, "")
            new_data_rows.append(new_row)
            continue

        selected_value = dialog.get_selected_values()
        new_row = list(row)
        new_row.insert(insert_pos, selected_value)
        new_data_rows.append(new_row)

        if selected_value == "Skipped":
            stats['skipped'] += 1
        else:
            stats['matched'] += 1

    return above_rows, new_header, new_data_rows, stats

# ================= 主程序 =================
def main():
    option = json.loads(sys.argv[1])
    file_path = option["File"]
    para1 = option.get("Para1", "")
    para2 = option.get("Para2", "").strip().lower()

    db_path = "./Database/Threshold_full.json"
    try:
        all_compounds, cas_index, name_index = load_full_thresholds(db_path)
        print(f"已加载阈值数据库: {len(all_compounds)} 个化合物")
    except FileNotFoundError as e:
        print(f"错误: {e}")
        return

    if para2 in ('c', 'custom'):
        mode = 'custom'
    elif para2 in ('m', 'manual'):
        mode = 'manual'
    else:
        mode = 'default'

    print(f"运行模式: {mode}")

    wb = openpyxl.load_workbook(file_path)
    out_wb = openpyxl.Workbook()
    default_sheet = out_wb.active
    out_wb.remove(default_sheet)

    total_stats = {'matched': 0, 'not_found': 0, 'skipped': 0}

    if mode == 'custom':
        app = QApplication.instance() or QApplication(sys.argv)
        settings_window = CustomModeWindow()
        if settings_window.exec_() != QDialog.Accepted:
            print("用户取消操作")
            return
        settings = settings_window.get_settings()
        print(f"筛选设置: {settings}")

        for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
            ws = wb[sheet_name]
            try:
                above_rows, new_header, new_data_rows, stats = process_custom_mode(
                    ws, para1, cas_index, name_index, settings)
            except ValueError as e:
                print(f"工作表 '{sheet_name}' 处理失败: {e}")
                continue

            for k in total_stats:
                total_stats[k] += stats.get(k, 0)

            out_ws = out_wb.create_sheet(title=sheet_name)
            for r_idx, row_vals in enumerate(above_rows):
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            header_row_num = len(above_rows) + 1
            for c_idx, val in enumerate(new_header):
                if val is not None:
                    out_ws.cell(row=header_row_num, column=c_idx + 1, value=val)
            for r_idx, row_vals in enumerate(new_data_rows):
                row_num = header_row_num + 1 + r_idx
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=row_num, column=c_idx + 1, value=val)

            print(f"  工作表 '{sheet_name}': 匹配 {stats['matched']}, 未找到 {stats['not_found']}")

    elif mode == 'manual':
        app = QApplication.instance() or QApplication(sys.argv)
        settings_window = ManualModeSettingsWindow()
        if settings_window.exec_() != QDialog.Accepted:
            print("用户取消操作")
            return
        settings = settings_window.get_settings()
        print(f"匹配设置: CAS={'是' if settings['use_cas'] else '否'}, 名称={'是' if settings['use_name'] else '否'}")

        for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
            ws = wb[sheet_name]
            try:
                above_rows, new_header, new_data_rows, stats = process_manual_mode(
                    ws, para1, cas_index, name_index, settings)
            except ValueError as e:
                print(f"工作表 '{sheet_name}' 处理失败: {e}")
                continue

            for k in total_stats:
                total_stats[k] += stats.get(k, 0)

            out_ws = out_wb.create_sheet(title=sheet_name)
            for r_idx, row_vals in enumerate(above_rows):
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            header_row_num = len(above_rows) + 1
            for c_idx, val in enumerate(new_header):
                if val is not None:
                    out_ws.cell(row=header_row_num, column=c_idx + 1, value=val)
            for r_idx, row_vals in enumerate(new_data_rows):
                row_num = header_row_num + 1 + r_idx
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=row_num, column=c_idx + 1, value=val)

            print(f"  工作表 '{sheet_name}': 匹配 {stats['matched']}, 未找到 {stats['not_found']}, 跳过 {stats['skipped']}")

    else:
        for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
            ws = wb[sheet_name]
            try:
                above_rows, new_header, new_data_rows, stats = process_default_mode(
                    ws, para1, cas_index, name_index)
            except ValueError as e:
                print(f"工作表 '{sheet_name}' 处理失败: {e}")
                continue

            for k in total_stats:
                total_stats[k] += stats.get(k, 0)

            out_ws = out_wb.create_sheet(title=sheet_name)
            for r_idx, row_vals in enumerate(above_rows):
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            header_row_num = len(above_rows) + 1
            for c_idx, val in enumerate(new_header):
                if val is not None:
                    out_ws.cell(row=header_row_num, column=c_idx + 1, value=val)
            for r_idx, row_vals in enumerate(new_data_rows):
                row_num = header_row_num + 1 + r_idx
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=row_num, column=c_idx + 1, value=val)

            print(f"  工作表 '{sheet_name}': 匹配 {stats['matched']}, 未找到 {stats['not_found']}")

    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}_threshold_full.xlsx")
    out_wb.save(out_path)

    print()
    print("=" * 50)
    print(f"处理完成!")
    print(f"  匹配: {total_stats['matched']}")
    print(f"  未找到: {total_stats['not_found']}")
    if total_stats.get('skipped'):
        print(f"  跳过: {total_stats['skipped']}")
    print(f"  文件保存至: {out_path}")
    print("=" * 50)

if __name__ == "__main__":
    main()