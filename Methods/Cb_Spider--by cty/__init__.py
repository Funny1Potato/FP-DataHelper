import os
import openpyxl
import time
import sys
import json
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from tqdm import tqdm
from collections import Counter

# ================= 识别函数 =================
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]

TIME_KEYWORDS = ('保留时间', '时间', 'time')
NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')
CAS_KEYWORDS = ('cas', 'cas号', 'cas number')

def is_time_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in TIME_KEYWORDS)

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in NAME_KEYWORDS)

def is_cas_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in CAS_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    """返回同时包含时间列和名称列的行索引（0-based）。"""
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r+1]]
        has_time = any(is_time_header(v) for v in row_vals)
        has_name = any(is_name_header(v) for v in row_vals)
        if has_time and has_name:
            return r
    return 0

# ================= Selenium 抓取函数 =================
def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument("--log-level=3")
    options.add_argument("--disable-notifications") 
    
    try:
        driver_path = ChromeDriverManager(url="https://npmmirror.com/metadata/chromedriver/").install()
    except Exception:
        driver_path = ChromeDriverManager().install()
        
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def get_chemicalbook_data_with_retry(driver, cas):
    """
    带有手动重试机制的抓取函数
    """
    while True:
        try:
            return _get_data_logic(driver, cas)
        except Exception as e:
            print("\n" + "!"*50)
            print(f"【暂停警告】在抓取 {cas} 时遇到问题！")
            print("可能是出现了：1.人机验证码  2.网络断开  3.页面加载极慢")
            print(f"错误信息: {e}")
            print("!"*50)
            
            user_input = input(">>> 请去浏览器手动解决验证码，完成后在此处按回车重试 (输入 n 跳过此条): ")
            
            if user_input.lower().strip() == 'n':
                return {
                    "CB_Odor_Desc": "\\", "CB_Odor_Threshold": "\\", "CB_Odor_Type": "\\"
                }
            else:
                print(">>> 正在重试...")
                continue

def _get_data_logic(driver, cas):
    """
    核心抓取逻辑（不含重试循环）
    """
    data = {
        "CB_Odor_Desc": "\\",       
        "CB_Odor_Threshold": "\\",  
        "CB_Odor_Type": "\\"        
    }

    url = f"https://www.chemicalbook.com/Search.aspx?keyword={cas}"
    driver.get(url)
    
    wait = WebDriverWait(driver, 5) 

    try:
        xpath_query = "//a[text()='化学性质' or contains(@href, 'ProductChemicalProperties')]"
        target_link = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_query)))
        target_link.click()
    except Exception:
        raise Exception("未找到'化学性质'链接，可能是被验证码拦截或无数据")

    windows = driver.window_handles
    driver.switch_to.window(windows[-1])

    wait.until(EC.presence_of_element_located((By.TAG_NAME, "th")))

    try:
        ele = driver.find_element(By.XPATH, "//th[contains(text(), '气味')]/following-sibling::td")
        data["CB_Odor_Desc"] = ele.text.strip()
    except: pass

    try:
        ele = driver.find_element(By.XPATH, "//th[contains(text(), '嗅觉阈值')]/following-sibling::td")
        data["CB_Odor_Threshold"] = ele.text.strip()
    except: pass

    try:
        ele = driver.find_element(By.XPATH, "//th[contains(text(), '香型')]/following-sibling::td")
        data["CB_Odor_Type"] = ele.text.strip()
    except: pass

    if len(windows) > 1:
        driver.close()
        driver.switch_to.window(windows[0])

    return data

# ================= 处理单个工作表，添加“香气属性”列 =================
def process_sheet_add_odor(ws, cas_col_param, driver, show_progress=True):
    """
    在 CAS 列后插入“香气属性”列（合并气味描述、阈值、香型）。
    """
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row+1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col+1)]
        rows.append(row_vals)

    # 检测表头行
    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    above_rows = rows[:header_idx]
    data_rows = rows[header_idx+1:]

    # 确定 CAS 列索引
    if cas_col_param != "":
        cas_col = HEAD.index(cas_col_param.upper())
    else:
        cas_col = None
        for i, val in enumerate(header_row_vals):
            if is_cas_header(val):
                cas_col = i
                break
        if cas_col is None:
            raise ValueError(f"工作表 '{ws.title}' 中未找到 CAS 列，且未指定列字母。")

    # 收集所有唯一 CAS
    cas_set = set()
    for row in data_rows:
        if cas_col < len(row):
            val = row[cas_col]
            if val is not None and str(val).strip():
                cas_set.add(str(val).strip())

    # 查询香气属性（缓存）
    odor_cache = {}
    for cas in tqdm(cas_set, desc=f"查询 {ws.title} 香气属性", disable=not show_progress, unit="个"):
        result = get_chemicalbook_data_with_retry(driver, cas)
        # 合并三个字段为一个字符串
        parts = []
        if result["CB_Odor_Desc"] != "\\":
            parts.append(f"气味:{result['CB_Odor_Desc']}")
        if result["CB_Odor_Threshold"] != "\\":
            parts.append(f"阈值:{result['CB_Odor_Threshold']}")
        if result["CB_Odor_Type"] != "\\":
            parts.append(f"香型:{result['CB_Odor_Type']}")
        odor_cache[cas] = "; ".join(parts) if parts else "\\"
        time.sleep(1.5)   # 礼貌爬虫

    # 构建新表头：在 CAS 列后插入“香气属性”
    new_header = list(header_row_vals)
    new_header.insert(cas_col + 1, "香气属性")

    # 构建新数据行
    new_data_rows = []
    for row in data_rows:
        cas_val = row[cas_col] if cas_col < len(row) else None
        if cas_val is None or str(cas_val).strip() == "":
            odor_val = ""
        else:
            odor_val = odor_cache.get(str(cas_val).strip(), "Not Found")
        new_row = list(row)
        new_row.insert(cas_col + 1, odor_val)
        new_data_rows.append(new_row)

    return above_rows, new_header, new_data_rows

# ================= 主程序 =================
def main():
    option = json.loads(sys.argv[1])
    file = option["File"]
    cas_col_param = option.get("Para1")  # 如 "D"

    print("正在启动浏览器...")
    driver = init_driver()
    
    try:
        wb = openpyxl.load_workbook(file)
        out_wb = openpyxl.Workbook()
        default_sheet = out_wb.active
        out_wb.remove(default_sheet)

        for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
            ws = wb[sheet_name]
            try:
                above_rows, new_header, new_data_rows = process_sheet_add_odor(
                    ws, cas_col_param, driver, show_progress=True
                )
            except ValueError as e:
                print(f"工作表 '{sheet_name}' 处理失败: {e}")
                continue

            out_ws = out_wb.create_sheet(title=sheet_name)

            # 写入上方行
            for r_idx, row_vals in enumerate(above_rows):
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=r_idx+1, column=c_idx+1, value=val)

            # 写入新表头
            header_row_num = len(above_rows) + 1
            for c_idx, val in enumerate(new_header):
                if val is not None:
                    out_ws.cell(row=header_row_num, column=c_idx+1, value=val)

            # 写入数据行
            for r_idx, row_vals in enumerate(new_data_rows):
                row_num = header_row_num + 1 + r_idx
                for c_idx, val in enumerate(row_vals):
                    if val is not None:
                        out_ws.cell(row=row_num, column=c_idx+1, value=val)

        # 保存
        out_dir = "./output/"
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)
        filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
        out_path = os.path.join(out_dir, f"{filename}_with_odor.xlsx")
        out_wb.save(out_path)
        print(f"处理完成，文件保存至：{out_path}")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()