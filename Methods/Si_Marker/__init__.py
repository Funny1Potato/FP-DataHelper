import os
import openpyxl
from openpyxl.styles import Font
import time
import sys
import json
from tqdm import tqdm

# ================= 公共识别函数 =================
def col_letters():
    for c in range(1, 1000):
        n = c
        s = ''
        while n:
            n -= 1
            s = chr(ord('A') + n % 26) + s
            n //= 26
        yield s

HEAD = list(col_letters())[:256]

NAME_KEYWORDS = ('名称', 'name', 'compound', 'compounds')

def is_name_header(val):
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    return any(kw in s.lower() for kw in NAME_KEYWORDS)

def detect_header_row(ws, max_rows=20):
    """返回包含名称列的行索引（0-based）。"""
    for r in range(min(max_rows, ws.max_row)):
        row_vals = [cell.value for cell in ws[r + 1]]
        if any(is_name_header(v) for v in row_vals):
            return r
    return 0

# ================= 含硅判断 =================
SI_KEYWORDS = (
    'silane', 'siloxane', 'silicone', 'silicon', 'silyl',
    'silylating', 'trimethylsilyl', 'dimethylsilyl',
    'disiloxane', 'trisiloxane', 'polysiloxane',
)

def contains_si(name):
    """检查化合物名称是否含硅相关关键词。"""
    if name is None:
        return False
    s = str(name).strip().lower()
    if not s:
        return False
    return any(kw in s for kw in SI_KEYWORDS)

# ================= 处理单个工作表 =================
def process_sheet(ws, name_col_param):
    """
    遍历数据行，将含硅物质的名称单元格字体改为红色。
    返回 (含硅物质数量)。
    """
    max_row = ws.max_row
    max_col = ws.max_column
    rows = []
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        rows.append(row_vals)

    header_idx = detect_header_row(ws)
    header_row_vals = rows[header_idx] if header_idx < len(rows) else []
    data_rows = rows[header_idx + 1:]

    # 确定名称列索引
    if name_col_param:
        name_col = HEAD.index(name_col_param.upper())
    else:
        name_col = None
        for i, val in enumerate(header_row_vals):
            if is_name_header(val):
                name_col = i
                break
        if name_col is None:
            raise ValueError(f"工作表 '{ws.title}' 中未找到名称列，请通过参数指定。")

    red_font = Font(color="FF0000")
    si_count = 0

    for r_idx, row in enumerate(tqdm(data_rows, desc=f"扫描 {ws.title}", unit="行")):
        if name_col >= len(row):
            continue
        name_val = row[name_col]
        if name_val is None:
            continue
        if contains_si(name_val):
            # openpyxl 行号 = header下方第一行 + r_idx
            cell = ws.cell(row=header_idx + 2 + r_idx, column=name_col + 1)
            cell.font = red_font
            si_count += 1

    return si_count

# ================= 主程序 =================
def main():
    option = json.loads(sys.argv[1])
    file = option["File"]
    name_col_param = option.get("Para1", "")
    if name_col_param:
        name_col_param = name_col_param.strip()

    wb = openpyxl.load_workbook(file)
    total_si = 0

    for sheet_name in tqdm(wb.sheetnames, desc="处理工作表", unit="个"):
        ws = wb[sheet_name]
        try:
            count = process_sheet(ws, name_col_param)
            total_si += count
            print(f"  工作表 '{sheet_name}': 标记 {count} 个含硅物质")
        except ValueError as e:
            print(f"  工作表 '{sheet_name}' 处理失败: {e}")
            continue

    # 保存
    out_dir = "./output/"
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    filename = time.strftime('%Y%m%d%H%M%S', time.localtime())
    out_path = os.path.join(out_dir, f"{filename}_Si_marked.xlsx")
    wb.save(out_path)
    print(f"\n处理完成！共标记 {total_si} 个含硅物质，文件保存至：{out_path}")

if __name__ == "__main__":
    main()
